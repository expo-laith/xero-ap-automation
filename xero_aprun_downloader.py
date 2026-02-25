# xero_aprun_downloader.py  (flat category folders + numbered filenames)
import os
import json
import base64
import time
import datetime as dt
from typing import Dict, Any, List, Optional

import requests
from openpyxl import load_workbook

SECRETS_FILE = "xero_secrets.json"

# ---------------- Environment helpers ----------------
def _env(*names):
    """Return the first non-empty environment variable from the given names."""
    for name in names:
        val = os.environ.get(name, "").strip()
        if val:
            return val
    return None

# ---------------- OAuth ----------------
def load_secrets() -> Dict[str, str]:
    """Load secrets from environment variables with fallback to xero_secrets.json if it exists."""
    secrets = {}
    
    # Try environment variables first (with fallback names)
    secrets["client_id"] = _env("XERO_CLIENT_ID", "CLIENT_ID") or ""
    secrets["client_secret"] = _env("XERO_CLIENT_SECRET", "CLIENT_SECRET") or ""
    secrets["refresh_token"] = _env("XERO_REFRESH_TOKEN", "REFRESH_TOKEN") or ""
    secrets["tenant_id"] = _env("XERO_TENANT_ID", "TENANT_ID") or ""
    secrets["scopes"] = _env("XERO_SCOPES", "SCOPES") or "offline_access accounting.transactions accounting.attachments"
    secrets["redirect_uri"] = _env("XERO_REDIRECT_URI", "REDIRECT_URI") or "http://localhost:8080/callback"
    
    # If any required field is missing and xero_secrets.json exists, load from file
    if (not secrets["client_id"] or not secrets["client_secret"] or not secrets["refresh_token"]) and os.path.exists(SECRETS_FILE):
        with open(SECRETS_FILE, "r", encoding="utf-8") as f:
            file_secrets = json.load(f)
            secrets["client_id"] = secrets["client_id"] or file_secrets.get("client_id", "")
            secrets["client_secret"] = secrets["client_secret"] or file_secrets.get("client_secret", "")
            secrets["refresh_token"] = secrets["refresh_token"] or file_secrets.get("refresh_token", "")
            secrets["tenant_id"] = secrets["tenant_id"] or file_secrets.get("tenant_id", "")
            secrets["scopes"] = secrets["scopes"] or file_secrets.get("scopes", "offline_access accounting.transactions accounting.attachments")
            secrets["redirect_uri"] = secrets["redirect_uri"] or file_secrets.get("redirect_uri", "http://localhost:8080/callback")
    
    if not secrets["client_id"] or not secrets["client_secret"] or not secrets["refresh_token"]:
        raise ValueError("Missing required credentials: client_id, client_secret, and refresh_token must be provided via environment variables or xero_secrets.json")
    
    return secrets

def save_secrets(s: Dict[str, str]):
    """Save secrets back to xero_secrets.json if running locally (file exists)."""
    if os.path.exists(SECRETS_FILE):
        with open(SECRETS_FILE, "w", encoding="utf-8") as f:
            json.dump(s, f, indent=2)
        print(f"[info] Updated {SECRETS_FILE} with new refresh token")

def refresh_access_token(secrets: Dict[str, str]) -> str:
    """Refresh the access token using the correct OAuth2 flow with Basic auth."""
    client_id = secrets["client_id"]
    client_secret = secrets["client_secret"]
    refresh_token = secrets["refresh_token"]
    
    # Create Basic auth header
    credentials = f"{client_id}:{client_secret}"
    b64_credentials = base64.b64encode(credentials.encode()).decode()
    
    headers = {
        "Authorization": f"Basic {b64_credentials}",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    
    data = {
        "grant_type": "refresh_token",
        "refresh_token": refresh_token
    }
    
    r = requests.post(
        "https://identity.xero.com/connect/token",
        headers=headers,
        data=data,
        timeout=40,
    )
    
    # Debug output as required
    print("TOKEN STATUS:", r.status_code)
    print("TOKEN BODY:", r.text)
    
    r.raise_for_status()
    tok = r.json()
    
    # Xero rotates refresh tokens - handle persistence
    new_refresh_token = tok.get("refresh_token")
    if new_refresh_token and new_refresh_token != refresh_token:
        secrets["refresh_token"] = new_refresh_token
        
        # Check if running in Azure (XERO_REFRESH_TOKEN env var present)
        if _env("XERO_REFRESH_TOKEN", "REFRESH_TOKEN"):
            print("=" * 80)
            print("WARNING: Refresh token has been rotated by Xero!")
            print("You are running in Azure App Service (environment variable detected).")
            print("You MUST manually update the XERO_REFRESH_TOKEN in Azure App Settings.")
            print("New refresh token:")
            print(new_refresh_token)
            print("=" * 80)
        else:
            # Running locally - save to file if it exists
            save_secrets(secrets)
    
    return tok["access_token"]

def get_tenant_id(token: str, secrets: Dict[str, str]) -> str:
    """Get tenant ID from env var or by querying Xero connections API."""
    tenant_id = secrets.get("tenant_id")
    if tenant_id:
        print(f"[info] Using tenant ID from configuration: {tenant_id}")
        return tenant_id
    
    # Query connections API
    print("[info] Tenant ID not configured, querying Xero connections...")
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    
    r = requests.get("https://api.xero.com/connections", headers=headers, timeout=30)
    r.raise_for_status()
    
    connections = r.json()
    if not connections:
        raise ValueError("No Xero connections found for this account")
    
    conn = connections[0]
    tenant_id = conn["tenantId"]
    tenant_name = conn.get("tenantName", "Unknown")
    
    print(f"[info] Selected tenant: {tenant_name} (ID: {tenant_id})")
    
    # Save for future use
    secrets["tenant_id"] = tenant_id
    save_secrets(secrets)
    
    return tenant_id

def xero_get(url: str, token: str, tenant_id: str, params=None, stream: bool=False):
    headers = {
        "Authorization": f"Bearer {token}",
        "Xero-tenant-id": tenant_id,
        "Accept": "*/*" if stream else "application/json",
    }
    for i in range(6):
        resp = requests.get(url, headers=headers, params=params, timeout=60, stream=stream)
        if resp.status_code in (429, 500, 502, 503, 504):
            backoff = min(60, 2**i)
            print(f"[warn] {resp.status_code} from Xero. Backing off {backoff}s…")
            time.sleep(backoff)
            continue
        resp.raise_for_status()
        return resp
    resp.raise_for_status()

# ---------------- Xero helpers ----------------
def query_invoices(token: str, tenant_id: str, where: str, page: int=1) -> List[Dict[str, Any]]:
    base = "https://api.xero.com/api.xro/2.0/Invoices"
    params = {"where": where, "order": "Date DESC", "page": page}
    resp = xero_get(base, token, tenant_id, params=params)
    return resp.json().get("Invoices", [])

def find_by_invoice_number(token: str, tenant_id: str, supplier: str, invnum_raw: str) -> Optional[Dict[str, Any]]:
    """Search by InvoiceNumber (where your PO-like strings are)."""
    variants = []
    invnum = (invnum_raw or "").strip()
    if invnum:
        variants += [invnum, invnum.replace(" ", ""), invnum.replace("/", "")]

    # exact with supplier, then without
    for v in dict.fromkeys(variants):
        where = f'Type=="ACCPAY" && InvoiceNumber=="{v}"'
        if supplier:
            where += f' && Contact.Name=="{supplier}"'
        hits = query_invoices(token, tenant_id, where, page=1)
        if hits:
            return hits[0]
    for v in dict.fromkeys(variants):
        where = f'Type=="ACCPAY" && InvoiceNumber=="{v}"'
        hits = query_invoices(token, tenant_id, where, page=1)
        if hits:
            return hits[0]

    # fallback: paginate & compare locally
    since = dt.datetime.utcnow() - dt.timedelta(days=365 if supplier else 120)
    where = f'Date>=DateTime({since.year},{since.month},{since.day},0,0,0) && Type=="ACCPAY"'
    if supplier:
        where += f' && Contact.Name=="{supplier}"'

    want_norms = { invnum.replace(" ", "").replace("/", "") }
    page = 1
    while page <= 10:
        hits = query_invoices(token, tenant_id, where, page=page)
        if not hits:
            break
        for inv in hits:
            num = (inv.get("InvoiceNumber") or "")
            if num.replace(" ", "").replace("/", "") in want_norms:
                return inv
        page += 1
    return None

def list_attachments(token: str, tenant_id: str, inv_id: str) -> List[Dict[str, Any]]:
    url = f"https://api.xero.com/api.xro/2.0/Invoices/{inv_id}/Attachments"
    return xero_get(url, token, tenant_id).json().get("Attachments", [])

def download_attachment(token: str, tenant_id: str, inv_id: str, filename: str) -> bytes:
    from urllib.parse import quote
    url = f"https://api.xero.com/api.xro/2.0/Invoices/{inv_id}/Attachments/{quote(filename)}"
    return xero_get(url, token, tenant_id, stream=True).content

# ---------------- Excel & IO ----------------
def safe(s: str) -> str:
    s = (s or "Unknown").strip()
    for ch in r'\/:*?"<>|':
        s = s.replace(ch, "-")
    return s

def read_aprun_rows(xlsx_path: str) -> List[Dict[str, str]]:
    """
    Accept either:
      - contact (supplier), invoice reference (invoice number), category
      - supplier, reference, category
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column+1):
        v = ws.cell(row=1, column=c).value
        key = (str(v).strip().lower() if v is not None else "")
        if key:
            headers[key] = c

    supplier_col = headers.get("supplier") or headers.get("contact")
    reference_col = headers.get("reference") or headers.get("invoice reference")
    category_col  = headers.get("category")

    if not (supplier_col and reference_col and category_col):
        raise ValueError(
            "Sheet must have headers including 'category' and either "
            "('supplier' or 'contact') and either ('reference' or 'invoice reference'). "
            f"Found: {list(headers.keys())}"
        )

    rows: List[Dict[str, str]] = []
    for r in range(2, ws.max_row+1):
        supplier = ws.cell(r, supplier_col).value
        invref   = ws.cell(r, reference_col).value
        category = ws.cell(r, category_col).value

        supplier = (str(supplier).strip() if supplier is not None else "")
        invref   = (str(invref).strip() if invref is not None else "")
        category = (str(category).strip() if category is not None else "")

        if not invref or not category:
            continue
        rows.append({"supplier": supplier, "invnum": invref, "category": category})
    return rows

def unique_path(base_dir: str, fname: str) -> str:
    """Avoid overwriting when same filename appears again."""
    path = os.path.join(base_dir, fname)
    if not os.path.exists(path):
        return path
    name, ext = os.path.splitext(fname)
    n = 2
    while True:
        cand = os.path.join(base_dir, f"{name} ({n}){ext}")
        if not os.path.exists(cand):
            return cand
        n += 1

# ---------------- Main ----------------
def main():
    # Determine paths based on environment (local vs Azure)
    # Local defaults
    default_xlsx = "AP run.xlsx"
    default_out = "./output"
    
    # Azure detection and defaults
    if os.path.exists("/home/site/wwwroot"):
        # Running in Azure App Service
        default_out = "/home/site/wwwroot/output"
    
    xlsx_path = os.environ.get("XERO_APRUN_XLSX", default_xlsx)
    out_root  = os.environ.get("XERO_OUT_ROOT", default_out)

    secrets   = load_secrets()
    token     = refresh_access_token(secrets)
    tenant_id = get_tenant_id(token, secrets)

    today_folder = os.path.join(out_root, dt.date.today().isoformat())
    os.makedirs(today_folder, exist_ok=True)

    # Ensure the 4 category folders exist (names sanitized for Windows)
    canonical_categories = [
        "Billable Projects",
        "Factory overheads / Consumables",
        "Exhibit Central",
        "USD",
    ]
    cat_dirs = {c: os.path.join(today_folder, safe(c).replace(" / ", " - ")) for c in canonical_categories}
    for p in cat_dirs.values():
        os.makedirs(p, exist_ok=True)

    rows = read_aprun_rows(xlsx_path)
    print(f"[info] {len(rows)} rows after filtering. Output: {today_folder}")

    # per-category numbering (01, 02, …) in the order rows appear; only increment when we actually save files
    cat_counters: Dict[str, int] = {c: 0 for c in canonical_categories}
    processed_invoice_ids = set()
    downloaded_keys = set()

    new_files = 0
    missing   = 0

    for i, row in enumerate(rows, start=1):
        sup, invnum, cat = row["supplier"], row["invnum"], row["category"]
        # normalize cat label to one of our 4 folders (by simple case-insensitive match)
        target_cat = next((c for c in canonical_categories if c.lower() == cat.lower()), cat)
        dest_dir   = cat_dirs.get(target_cat, os.path.join(today_folder, safe(target_cat).replace(" / "," - ")))
        if dest_dir not in cat_dirs.values():
            # if an unexpected category appears, still create it flatly
            os.makedirs(dest_dir, exist_ok=True)

        print(f"[{i}] Lookup: {sup or '—'} | InvoiceNumber='{invnum}' | Category='{target_cat}'")
        inv = find_by_invoice_number(token, tenant_id, sup, invnum)
        if not inv:
            print("    ✗ NOT FOUND")
            missing += 1
            continue

        inv_id = inv["InvoiceID"]
        inv_no = inv.get("InvoiceNumber") or inv_id
        contact= ((inv.get("Contact") or {}).get("Name") or sup or "Unknown")

        if inv_id in processed_invoice_ids:
            print(f"Skipping invoice already processed: {inv_id}")
            continue

        atts = list_attachments(token, tenant_id, inv_id)
        if not atts:
            print("    (no attachments)")
            processed_invoice_ids.add(inv_id)
            continue

        # increment category counter ONCE per invoice we actually save attachments for
        cat_counters[target_cat] = cat_counters.get(target_cat, 0) + 1
        seq = cat_counters[target_cat]
        seq_tag = f"{seq:02d}"

        for a in atts:
            att_id = a.get("AttachmentID") or a.get("AttachmentId") or a.get("attachmentId")
            key = (inv_id, att_id or a["FileName"])
            if key in downloaded_keys:
                print(f"[skip] already downloaded: {a['FileName']}")
                continue
            orig = a["FileName"]
            filename = f"{seq_tag} - {safe(contact)} - {safe(inv_no)} - {orig}"
            outp = unique_path(dest_dir, filename)
            blob = download_attachment(token, tenant_id, inv_id, a["FileName"])
            with open(outp, "wb") as f:
                f.write(blob)
            downloaded_keys.add(key)
            new_files += 1
            print(f"    ✓ {os.path.basename(outp)}")

        processed_invoice_ids.add(inv_id)

    print(f"[done] New files: {new_files}. Missing invoices: {missing}. Output: {today_folder}")

if __name__ == "__main__":
    main()
