# xero_aprun_downloader.py  (flat category folders + numbered filenames)
import base64
import os
import json
import time
import datetime as dt
from typing import Dict, Any, List, Optional

import requests
from openpyxl import load_workbook

SECRETS_FILE = "xero_secrets.json"

# ---------------- OAuth ----------------
def load_secrets() -> Dict[str, str]:
    with open(SECRETS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_secrets(s: Dict[str, str]):
    with open(SECRETS_FILE, "w", encoding="utf-8") as f:
        json.dump(s, f, indent=2)

def refresh_access_token(secrets: Dict[str, str]) -> str:
    import base64

    basic_auth = base64.b64encode(
        f"{secrets['client_id']}:{secrets['client_secret']}".encode("utf-8")
    ).decode("ascii")

    headers = {
        "Authorization": f"Basic {basic_auth}",
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json",
    }

    data = {
        "grant_type": "refresh_token",
        "refresh_token": secrets["refresh_token"],
    }

    r = requests.post(
        "https://identity.xero.com/connect/token",
        headers=headers,
        data=data,
        timeout=40,
    )

    print("TOKEN STATUS:", r.status_code)
    print("TOKEN BODY:", r.text)

    r.raise_for_status()

    tok = r.json()

    # Xero rotates refresh tokens
    secrets["refresh_token"] = tok["refresh_token"]
    save_secrets(secrets)

    return tok["access_token"]

def xero_get(url: str, token: str, tenant_id: str, params=None, stream: bool=False):
    headers = {
        "Authorization": f"Bearer {token}",
        "Xero-tenant-id": tenant_id,
        "Accept": "*/*" if stream else "application/json",
    }
    for i in range(6):
        resp = requests.get(url, headers=headers, params=params, timeout=60, stream=stream)
        if resp.status_code in (429, 500, 502, 503, 504):
            backoff = min(60, 2**i)
            print(f"[warn] {resp.status_code} from Xero. Backing off {backoff}s…")
            time.sleep(backoff)
            continue
        resp.raise_for_status()
        return resp
    resp.raise_for_status()

# ---------------- Xero helpers ----------------
def query_invoices(token: str, tenant_id: str, where: str, page: int=1) -> List[Dict[str, Any]]:
    base = "https://api.xero.com/api.xro/2.0/Invoices"
    params = {"where": where, "order": "Date DESC", "page": page}
    resp = xero_get(base, token, tenant_id, params=params)
    return resp.json().get("Invoices", [])

def find_by_invoice_number(token: str, tenant_id: str, supplier: str, invnum_raw: str) -> Optional[Dict[str, Any]]:
    """Search by InvoiceNumber (where your PO-like strings are)."""
    variants = []
    invnum = (invnum_raw or "").strip()
    if invnum:
        variants += [invnum, invnum.replace(" ", ""), invnum.replace("/", "")]

    # exact with supplier, then without
    for v in dict.fromkeys(variants):
        where = f'Type=="ACCPAY" && InvoiceNumber=="{v}"'
        if supplier:
            where += f' && Contact.Name=="{supplier}"'
        hits = query_invoices(token, tenant_id, where, page=1)
        if hits:
            return hits[0]
    for v in dict.fromkeys(variants):
        where = f'Type=="ACCPAY" && InvoiceNumber=="{v}"'
        hits = query_invoices(token, tenant_id, where, page=1)
        if hits:
            return hits[0]

    # fallback: paginate & compare locally
    since = dt.datetime.utcnow() - dt.timedelta(days=365 if supplier else 120)
    where = f'Date>=DateTime({since.year},{since.month},{since.day},0,0,0) && Type=="ACCPAY"'
    if supplier:
        where += f' && Contact.Name=="{supplier}"'

    want_norms = { invnum.replace(" ", "").replace("/", "") }
    page = 1
    while page <= 10:
        hits = query_invoices(token, tenant_id, where, page=page)
        if not hits:
            break
        for inv in hits:
            num = (inv.get("InvoiceNumber") or "")
            if num.replace(" ", "").replace("/", "") in want_norms:
                return inv
        page += 1
    return None

def list_attachments(token: str, tenant_id: str, inv_id: str) -> List[Dict[str, Any]]:
    url = f"https://api.xero.com/api.xro/2.0/Invoices/{inv_id}/Attachments"
    return xero_get(url, token, tenant_id).json().get("Attachments", [])

def download_attachment(token: str, tenant_id: str, inv_id: str, filename: str) -> bytes:
    from urllib.parse import quote
    url = f"https://api.xero.com/api.xro/2.0/Invoices/{inv_id}/Attachments/{quote(filename)}"
    return xero_get(url, token, tenant_id, stream=True).content

# ---------------- Excel & IO ----------------
def safe(s: str) -> str:
    s = (s or "Unknown").strip()
    for ch in r'\/:*?"<>|':
        s = s.replace(ch, "-")
    return s

def read_aprun_rows(xlsx_path: str) -> List[Dict[str, str]]:
    """
    Accept either:
      - contact (supplier), invoice reference (invoice number), category
      - supplier, reference, category
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column+1):
        v = ws.cell(row=1, column=c).value
        key = (str(v).strip().lower() if v is not None else "")
        if key:
            headers[key] = c

    supplier_col = headers.get("supplier") or headers.get("contact")
    reference_col = headers.get("reference") or headers.get("invoice reference")
    category_col  = headers.get("category")

    if not (supplier_col and reference_col and category_col):
        raise ValueError(
            "Sheet must have headers including 'category' and either "
            "('supplier' or 'contact') and either ('reference' or 'invoice reference'). "
            f"Found: {list(headers.keys())}"
        )

    rows: List[Dict[str, str]] = []
    for r in range(2, ws.max_row+1):
        supplier = ws.cell(r, supplier_col).value
        invref   = ws.cell(r, reference_col).value
        category = ws.cell(r, category_col).value

        supplier = (str(supplier).strip() if supplier is not None else "")
        invref   = (str(invref).strip() if invref is not None else "")
        category = (str(category).strip() if category is not None else "")

        if not invref or not category:
            continue
        rows.append({"supplier": supplier, "invnum": invref, "category": category})
    return rows

def unique_path(base_dir: str, fname: str) -> str:
    """Avoid overwriting when same filename appears again."""
    path = os.path.join(base_dir, fname)
    if not os.path.exists(path):
        return path
    name, ext = os.path.splitext(fname)
    n = 2
    while True:
        cand = os.path.join(base_dir, f"{name} ({n}){ext}")
        if not os.path.exists(cand):
            return cand
        n += 1

# ---------------- Main ----------------
def main():
    # Adjust these two if needed
    xlsx_path = os.environ.get("XERO_APRUN_XLSX", r"C:\Users\laith\OneDrive\Desktop\Xero project 2\AP run.xlsx")
    out_root  = os.environ.get("XERO_OUT_ROOT",  r"C:\Users\laith\OneDrive\Desktop\Xero project 2")

    secrets   = load_secrets()
    token     = refresh_access_token(secrets)
    tenant_id = secrets["tenant_id"]

    today_folder = os.path.join(out_root, dt.date.today().isoformat())
    os.makedirs(today_folder, exist_ok=True)

    # Ensure the category folders exist (flat)
    canonical_categories = [
        "Billable Projects",
        "Factory overheads / Consumables",
        "Exhibit Central",
        "USD",
        "Industric",  # ← new optional category
    ]
    cat_dirs = {c: os.path.join(today_folder, safe(c).replace(" / ", " - ")) for c in canonical_categories}
    for p in cat_dirs.values():
        os.makedirs(p, exist_ok=True)

    rows = read_aprun_rows(xlsx_path)
    print(f"[info] {len(rows)} rows after filtering. Output: {today_folder}")

    # per-category numbering (01, 02, …). Increments only when we actually save files.
    cat_counters: Dict[str, int] = {c: 0 for c in canonical_categories}

    new_files = 0
    missing   = 0

    for i, row in enumerate(rows, start=1):
        sup, invnum, cat = row["supplier"], row["invnum"], row["category"]
        # normalize cat label to one of our folders by case-insensitive match
        target_cat = next((c for c in canonical_categories if c.lower() == cat.lower()), cat)
        dest_dir   = cat_dirs.get(target_cat, os.path.join(today_folder, safe(target_cat).replace(" / "," - ")))
        if dest_dir not in cat_dirs.values():
            os.makedirs(dest_dir, exist_ok=True)  # unexpected categories still get created flat

        print(f"[{i}] Lookup: {sup or '—'} | InvoiceNumber='{invnum}' | Category='{target_cat}'")
        inv = find_by_invoice_number(token, tenant_id, sup, invnum)
        if not inv:
            print("    ✗ NOT FOUND")
            missing += 1
            continue

        inv_id = inv["InvoiceID"]
        inv_no = inv.get("InvoiceNumber") or inv_id
        contact= ((inv.get("Contact") or {}).get("Name") or sup or "Unknown")

        atts = list_attachments(token, tenant_id, inv_id)
        if not atts:
            print("    (no attachments)")
            continue

        # increment this category's sequence
        cat_counters[target_cat] = cat_counters.get(target_cat, 0) + 1
        seq_tag = f"{cat_counters[target_cat]:02d}"

        for a in atts:
            orig = a["FileName"]
            filename = f"{seq_tag} - {safe(contact)} - {safe(inv_no)} - {orig}"
            outp = unique_path(dest_dir, filename)
            blob = download_attachment(token, tenant_id, inv_id, a["FileName"])
            with open(outp, "wb") as f:
                f.write(blob)
            new_files += 1
            print(f"    ✓ {os.path.basename(outp)}")

    print(f"[done] New files: {new_files}. Missing invoices: {missing}. Output: {today_folder}")

if __name__ == "__main__":
    main()
