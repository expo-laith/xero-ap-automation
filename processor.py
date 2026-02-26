# xero_aprun_downloader.py  (flat category folders + numbered filenames)
import base64
import os
import json
import time
import traceback
import datetime as dt
from typing import Dict, Any, List, Optional

import requests
from openpyxl import load_workbook

BASE_DIR = os.getcwd()
SECRETS_FILE = os.path.join(BASE_DIR, "xero_secrets.json")

# ---------------- OAuth ----------------
def load_secrets() -> Dict[str, str]:
    # For GitHub/Azure deployment we load credentials from environment variables
    # so secrets are not stored in source control.
    secrets: Dict[str, str] = {
        "client_id": os.environ["XERO_CLIENT_ID"],
        "client_secret": os.environ["XERO_CLIENT_SECRET"],
        "redirect_uri": os.environ["XERO_REDIRECT_URI"],
        "tenant_id": os.environ["XERO_TENANT_ID"],
    }

    refresh_token = os.environ.get("XERO_REFRESH_TOKEN")
    if refresh_token:
        secrets["refresh_token"] = refresh_token

    if os.path.exists(SECRETS_FILE):
        with open(SECRETS_FILE, "r", encoding="utf-8") as f:
            file_secrets = json.load(f)
        if isinstance(file_secrets, dict):
            # Prefer rotated token from local runtime file if present.
            if file_secrets.get("refresh_token"):
                secrets["refresh_token"] = file_secrets["refresh_token"]

    if "refresh_token" not in secrets:
        raise KeyError(
            "Missing XERO_REFRESH_TOKEN environment variable (or runtime xero_secrets.json refresh_token)."
        )

    return secrets

def save_secrets(s: Dict[str, str]):
    # Keep a local runtime cache file for rotated refresh tokens between runs.
    # This file is intentionally gitignored and should not be committed.
    existing: Dict[str, Any] = {}
    if os.path.exists(SECRETS_FILE):
        try:
            with open(SECRETS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    existing = data
        except Exception as e:
            print(f"[WARN] Failed to load existing secrets file: {e}")
            traceback.print_exc()
            existing = {}

    # Persist only runtime values needed for token rotation between runs.
    for key in ("refresh_token", "tenant_id", "access_token", "token_type", "expires_in", "expires_at", "obtained_at"):
        if key in s:
            existing[key] = s[key]

    try:
        with open(SECRETS_FILE, "w", encoding="utf-8") as f:
            json.dump(existing, f, indent=2)
        print(f"[INFO] Saved secrets to {SECRETS_FILE}")
    except Exception as e:
        print(f"[ERROR] Failed to save secrets file: {e}")
        traceback.print_exc()
        raise

def refresh_access_token(secrets: Dict[str, str]) -> str:
    import base64

    basic_auth = base64.b64encode(
        f"{secrets['client_id']}:{secrets['client_secret']}".encode("utf-8")
    ).decode("ascii")

    headers = {
        "Authorization": f"Basic {basic_auth}",
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json",
    }

    data = {
        "grant_type": "refresh_token",
        "refresh_token": secrets["refresh_token"],
    }

    r = requests.post(
        "https://identity.xero.com/connect/token",
        headers=headers,
        data=data,
        timeout=40,
    )

    print("TOKEN STATUS:", r.status_code)
    print("TOKEN BODY:", r.text)

    r.raise_for_status()

    tok = r.json()

    # Xero rotates refresh tokens
    secrets["refresh_token"] = tok["refresh_token"]
    save_secrets(secrets)

    return tok["access_token"]

def xero_get(url: str, token: str, tenant_id: str, params=None, stream: bool=False):
    headers = {
        "Authorization": f"Bearer {token}",
        "Xero-tenant-id": tenant_id,
        "Accept": "*/*" if stream else "application/json",
    }
    for i in range(6):
        resp = requests.get(url, headers=headers, params=params, timeout=60, stream=stream)
        if resp.status_code in (429, 500, 502, 503, 504):
            backoff = min(60, 2**i)
            print(f"[warn] {resp.status_code} from Xero. Backing off {backoff}s…")
            time.sleep(backoff)
            continue
        resp.raise_for_status()
        return resp
    resp.raise_for_status()

# ---------------- Xero helpers ----------------
def query_invoices(token: str, tenant_id: str, where: str, page: int=1) -> List[Dict[str, Any]]:
    base = "https://api.xero.com/api.xro/2.0/Invoices"
    params = {"where": where, "order": "Date DESC", "page": page}
    resp = xero_get(base, token, tenant_id, params=params)
    return resp.json().get("Invoices", [])

def find_by_invoice_number(token: str, tenant_id: str, supplier: str, invnum_raw: str) -> Optional[Dict[str, Any]]:
    """Search by InvoiceNumber (where your PO-like strings are)."""
    variants = []
    invnum = (invnum_raw or "").strip()
    if invnum:
        variants += [invnum, invnum.replace(" ", ""), invnum.replace("/", "")]

    # exact with supplier, then without
    for v in dict.fromkeys(variants):
        where = f'Type=="ACCPAY" && InvoiceNumber=="{v}"'
        if supplier:
            where += f' && Contact.Name=="{supplier}"'
        hits = query_invoices(token, tenant_id, where, page=1)
        if hits:
            return hits[0]
    for v in dict.fromkeys(variants):
        where = f'Type=="ACCPAY" && InvoiceNumber=="{v}"'
        hits = query_invoices(token, tenant_id, where, page=1)
        if hits:
            return hits[0]

    # fallback: paginate & compare locally
    since = dt.datetime.utcnow() - dt.timedelta(days=365 if supplier else 120)
    where = f'Date>=DateTime({since.year},{since.month},{since.day},0,0,0) && Type=="ACCPAY"'
    if supplier:
        where += f' && Contact.Name=="{supplier}"'

    want_norms = { invnum.replace(" ", "").replace("/", "") }
    page = 1
    while page <= 10:
        hits = query_invoices(token, tenant_id, where, page=page)
        if not hits:
            break
        for inv in hits:
            num = (inv.get("InvoiceNumber") or "")
            if num.replace(" ", "").replace("/", "") in want_norms:
                return inv
        page += 1
    return None

def list_attachments(token: str, tenant_id: str, inv_id: str) -> List[Dict[str, Any]]:
    url = f"https://api.xero.com/api.xro/2.0/Invoices/{inv_id}/Attachments"
    return xero_get(url, token, tenant_id).json().get("Attachments", [])

def download_attachment(token: str, tenant_id: str, inv_id: str, filename: str) -> bytes:
    from urllib.parse import quote
    url = f"https://api.xero.com/api.xro/2.0/Invoices/{inv_id}/Attachments/{quote(filename)}"
    return xero_get(url, token, tenant_id, stream=True).content

# ---------------- Excel & IO ----------------
def safe(s: str) -> str:
    s = (s or "Unknown").strip()
    for ch in r'\/:*?"<>|':
        s = s.replace(ch, "-")
    return s

def read_aprun_rows(xlsx_path: str) -> List[Dict[str, str]]:
    """
    Accept either:
      - contact (supplier), invoice reference (invoice number), category
      - supplier, reference, category
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column+1):
        v = ws.cell(row=1, column=c).value
        key = (str(v).strip().lower() if v is not None else "")
        if key:
            headers[key] = c

    supplier_col = headers.get("supplier") or headers.get("contact")
    reference_col = headers.get("reference") or headers.get("invoice reference")
    category_col  = headers.get("category")

    if not (supplier_col and reference_col and category_col):
        raise ValueError(
            "Sheet must have headers including 'category' and either "
            "('supplier' or 'contact') and either ('reference' or 'invoice reference'). "
            f"Found: {list(headers.keys())}"
        )

    rows: List[Dict[str, str]] = []
    for r in range(2, ws.max_row+1):
        supplier = ws.cell(r, supplier_col).value
        invref   = ws.cell(r, reference_col).value
        category = ws.cell(r, category_col).value

        supplier = (str(supplier).strip() if supplier is not None else "")
        invref   = (str(invref).strip() if invref is not None else "")
        category = (str(category).strip() if category is not None else "")

        if not invref or not category:
            continue
        rows.append({"supplier": supplier, "invnum": invref, "category": category})
    return rows

def unique_path(base_dir: str, fname: str) -> str:
    """Avoid overwriting when same filename appears again."""
    path = os.path.join(base_dir, fname)
    if not os.path.exists(path):
        return path
    name, ext = os.path.splitext(fname)
    n = 2
    while True:
        cand = os.path.join(base_dir, f"{name} ({n}){ext}")
        if not os.path.exists(cand):
            return cand
        n += 1

# ---------------- Main ----------------
def run_ap_process(xlsx_path: str):
    try:
        print(f"\n[INFO] Starting AP process with file: {xlsx_path}")
        out_root  = os.environ.get("XERO_OUT_ROOT", BASE_DIR)
        print(f"[INFO] Output root: {out_root}")

        print("[INFO] Loading secrets...")
        secrets   = load_secrets()
        print("[INFO] Refreshing access token...")
        token     = refresh_access_token(secrets)
        tenant_id = secrets["tenant_id"]
        print(f"[INFO] Using tenant ID: {tenant_id}")

        today_folder = os.path.join(out_root, dt.date.today().isoformat())
        os.makedirs(today_folder, exist_ok=True)
        print(f"[INFO] Created output folder: {today_folder}")

        # Ensure the category folders exist (flat)
        canonical_categories = [
            "Billable Projects",
            "Factory overheads / Consumables",
            "Exhibit Central",
            "USD",
            "Industric",  # ← new optional category
        ]
        cat_dirs = {c: os.path.join(today_folder, safe(c).replace(" / ", " - ")) for c in canonical_categories}
        for p in cat_dirs.values():
            os.makedirs(p, exist_ok=True)

        print("[INFO] Reading AP run rows from Excel...")
        rows = read_aprun_rows(xlsx_path)
        print(f"[INFO] {len(rows)} rows after filtering. Output: {today_folder}")

        # per-category numbering (01, 02, …). Increments only when we actually save files.
        cat_counters: Dict[str, int] = {c: 0 for c in canonical_categories}
        processed_invoice_ids = set()
        downloaded_keys = set()

        new_files = 0
        missing   = 0

        for i, row in enumerate(rows, start=1):
            sup, invnum, cat = row["supplier"], row["invnum"], row["category"]
            # normalize cat label to one of our folders by case-insensitive match
            target_cat = next((c for c in canonical_categories if c.lower() == cat.lower()), cat)
            dest_dir   = cat_dirs.get(target_cat, os.path.join(today_folder, safe(target_cat).replace(" / "," - ")))
            if dest_dir not in cat_dirs.values():
                os.makedirs(dest_dir, exist_ok=True)  # unexpected categories still get created flat

            print(f"[{i}] Lookup: {sup or '—'} | InvoiceNumber='{invnum}' | Category='{target_cat}'")
            inv = find_by_invoice_number(token, tenant_id, sup, invnum)
            if not inv:
                print("    ✗ NOT FOUND")
                missing += 1
                continue

            inv_id = inv["InvoiceID"]
            inv_no = inv.get("InvoiceNumber") or inv_id
            contact= ((inv.get("Contact") or {}).get("Name") or sup or "Unknown")

            if inv_id in processed_invoice_ids:
                print(f"Skipping invoice already processed: {inv_id}")
                continue

            atts = list_attachments(token, tenant_id, inv_id)
            if not atts:
                print("    (no attachments)")
                processed_invoice_ids.add(inv_id)
                continue

            # increment this category's sequence
            cat_counters[target_cat] = cat_counters.get(target_cat, 0) + 1
            seq_tag = f"{cat_counters[target_cat]:02d}"

            for a in atts:
                att_id = a.get("AttachmentID") or a.get("AttachmentId") or a.get("attachmentId")
                key = (inv_id, att_id or a["FileName"])
                if key in downloaded_keys:
                    print(f"[skip] already downloaded: {a['FileName']}")
                    continue
                orig = a["FileName"]
                filename = f"{seq_tag} - {safe(contact)} - {safe(inv_no)} - {orig}"
                outp = unique_path(dest_dir, filename)
                blob = download_attachment(token, tenant_id, inv_id, a["FileName"])
                with open(outp, "wb") as f:
                    f.write(blob)
                downloaded_keys.add(key)
                new_files += 1
                print(f"    [ok] {os.path.basename(outp)}")

            processed_invoice_ids.add(inv_id)

        print(f"[done] New files: {new_files}. Missing invoices: {missing}. Output: {today_folder}")
        return {
            "new_files": new_files,
            "missing": missing,
            "output_folder": today_folder
        }
    except Exception as e:
        print("\n" + "="*80)
        print("ERROR IN run_ap_process")
        print("="*80)
        print(f"Exception type: {type(e).__name__}")
        print(f"Exception message: {str(e)}")
        print("\nFull traceback:")
        traceback.print_exc()
        print("="*80 + "\n")
        raise
